"""Report generation service for exporting analysis results to PDF and Excel."""

import io
from typing import List, Dict, Optional, Any
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, PageTemplate, Frame
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..models.contract import Contract
from ..models.clause import Clause
from ..models.finding import LeakageFinding, Severity
from ..db import get_cosmos_client, ContractRepository, ClauseRepository, FindingRepository
from ..utils.logging import setup_logging
from ..utils.exceptions import ReportGenerationError
from ..utils.brand_constants import BrandColors, Typography, Layout, ReportConfig

logger = setup_logging(__name__)


class ReportService:
    """Service for generating analysis reports in PDF and Excel formats."""

    def __init__(self):
        """Initialize report service."""
        logger.info("Report service initialized")

    def generate_pdf_report(
        self,
        contract_id: str,
        include_clauses: bool = False
    ) -> bytes:
        """
        Generate PDF report for contract analysis.

        Args:
            contract_id: Contract identifier
            include_clauses: Include full clause text in report

        Returns:
            PDF file content as bytes

        Raises:
            ReportGenerationError: If report generation fails
        """
        try:
            logger.info(f"Generating PDF report for contract {contract_id}")

            # Get data
            contract, clauses, findings = self._get_report_data(contract_id)

            # Create PDF in memory
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                topMargin=Layout.MARGIN_TOP,
                bottomMargin=Layout.MARGIN_BOTTOM,
                leftMargin=Layout.MARGIN_LEFT,
                rightMargin=Layout.MARGIN_RIGHT
            )

            # Build document content
            story = []

            # Add professional cover page
            story.extend(self._build_cover_page(contract, findings))
            story.append(PageBreak())

            # Add executive summary
            story.extend(self._build_executive_summary(contract, findings))
            story.append(PageBreak())

            # Add findings section
            story.extend(self._build_findings_section(findings, styles))

            if include_clauses and clauses:
                story.append(PageBreak())
                story.extend(self._build_clauses_section(clauses, styles))

            # Build PDF
            doc.build(story)

            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()

            logger.info(f"PDF report generated: {len(pdf_bytes)} bytes")

            return pdf_bytes

        except Exception as e:
            logger.error(f"PDF generation failed: {str(e)}")
            raise ReportGenerationError(f"PDF generation failed: {str(e)}")

    def generate_excel_report(self, contract_id: str) -> bytes:
        """
        Generate Excel report for contract analysis.

        Args:
            contract_id: Contract identifier

        Returns:
            Excel file content as bytes

        Raises:
            ReportGenerationError: If report generation fails
        """
        try:
            logger.info(f"Generating Excel report for contract {contract_id}")

            # Get data
            contract, clauses, findings = self._get_report_data(contract_id)

            # Create workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Add worksheets
            self._create_summary_sheet(wb, contract, findings)
            self._create_findings_sheet(wb, findings)
            self._create_clauses_sheet(wb, clauses)

            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            logger.info(f"Excel report generated: {len(excel_bytes)} bytes")

            return excel_bytes

        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise ReportGenerationError(f"Excel generation failed: {str(e)}")

    def _get_report_data(
        self,
        contract_id: str
    ) -> tuple[Contract, List[Clause], List[LeakageFinding]]:
        """Get all data needed for report generation."""
        cosmos_client = get_cosmos_client()

        # Get contract
        contract_repo = ContractRepository(cosmos_client.contracts_container)
        contract = contract_repo.get_by_contract_id(contract_id)

        if not contract:
            raise ReportGenerationError(f"Contract not found: {contract_id}")

        # Get clauses
        clause_repo = ClauseRepository(cosmos_client.clauses_container)
        clauses = clause_repo.get_by_contract_id(contract_id)

        # Get findings
        finding_repo = FindingRepository(cosmos_client.findings_container)
        findings = finding_repo.get_by_contract_id(contract_id)

        return contract, clauses, findings

    def _build_cover_page(
        self,
        contract: Contract,
        findings: List[LeakageFinding]
    ) -> List:
        """Build professional cover page inspired by KPMG Master Guide."""
        story = []

        # Title with brand color background
        title_style = ParagraphStyle(
            'CoverTitle',
            fontName=Typography.PRIMARY_BOLD,
            fontSize=Typography.COVER_TITLE,
            textColor=BrandColors.WHITE,
            alignment=TA_CENTER,
            spaceAfter=Layout.PARAGRAPH_SPACING
        )

        subtitle_style = ParagraphStyle(
            'CoverSubtitle',
            fontName=Typography.PRIMARY_BOLD,
            fontSize=Typography.COVER_SUBTITLE,
            textColor=BrandColors.WHITE,
            alignment=TA_CENTER,
            spaceAfter=Layout.PARAGRAPH_SPACING
        )

        # Create colored background block with title
        cover_data = [[
            Paragraph(ReportConfig.REPORT_TITLE, title_style)
        ]]

        cover_table = Table(cover_data, colWidths=[6*inch])
        cover_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BrandColors.PRIMARY_BLUE),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 50),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 50),
        ]))

        story.append(Spacer(1, 2*inch))  # Top spacing
        story.append(cover_table)
        story.append(Spacer(1, 0.3*inch))

        # Contract name
        contract_name_style = ParagraphStyle(
            'ContractName',
            fontName=Typography.PRIMARY_BOLD,
            fontSize=Typography.HEADING_1,
            textColor=BrandColors.DARK_GREY,
            alignment=TA_CENTER,
            spaceAfter=Layout.PARAGRAPH_SPACING
        )

        story.append(Paragraph(f"<b>{contract.contract_name}</b>", contract_name_style))
        story.append(Spacer(1, 0.5*inch))

        # Key metrics in colored boxes
        metrics_style = ParagraphStyle(
            'Metrics',
            fontName=Typography.PRIMARY_FONT,
            fontSize=Typography.BODY_LARGE,
            textColor=BrandColors.DARK_GREY,
            alignment=TA_CENTER
        )

        # Count findings by severity
        severity_counts = self._count_by_severity(findings)
        critical_count = severity_counts.get(Severity.CRITICAL, 0)
        high_count = severity_counts.get(Severity.HIGH, 0)
        total_findings = len(findings)

        metrics_data = [
            [
                Paragraph(f"<b>{total_findings}</b><br/>Total Findings", metrics_style),
                Paragraph(f"<b>{critical_count}</b><br/>Critical Issues", metrics_style),
                Paragraph(f"<b>{high_count}</b><br/>High Priority", metrics_style)
            ]
        ]

        metrics_table = Table(metrics_data, colWidths=[2*inch, 2*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), BrandColors.ACCENT_BLUE),
            ('BACKGROUND', (1, 0), (1, 0), BrandColors.CRITICAL_RED),
            ('BACKGROUND', (2, 0), (2, 0), BrandColors.HIGH_ORANGE),
            ('TEXTCOLOR', (0, 0), (-1, -1), BrandColors.WHITE),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 20),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
            ('GRID', (0, 0), (-1, -1), 0.5, BrandColors.WHITE),
        ]))

        story.append(metrics_table)
        story.append(Spacer(1, 1*inch))

        # Report metadata at bottom
        metadata_style = ParagraphStyle(
            'Metadata',
            fontName=Typography.PRIMARY_FONT,
            fontSize=Typography.BODY_SMALL,
            textColor=BrandColors.MEDIUM_GREY,
            alignment=TA_CENTER
        )

        generation_date = datetime.now().strftime('%B %d, %Y')
        story.append(Paragraph(f"Report Generated: {generation_date}", metadata_style))
        story.append(Paragraph(f"Contract ID: {contract.id}", metadata_style))
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(ReportConfig.ORGANIZATION_NAME, metadata_style))

        return story

    def _build_executive_summary(
        self,
        contract: Contract,
        findings: List[LeakageFinding]
    ) -> List:
        """Build executive summary section for PDF."""
        story = []

        # Section title with brand styling
        heading_style = ParagraphStyle(
            'SectionHeading',
            fontName=Typography.PRIMARY_BOLD,
            fontSize=Typography.SECTION_TITLE,
            textColor=BrandColors.PRIMARY_BLUE,
            spaceAfter=Layout.PARAGRAPH_SPACING,
            spaceBefore=Layout.SECTION_SPACING
        )

        story.append(Paragraph("Executive Summary", heading_style))
        story.append(Spacer(1, 0.2*inch))

        # Contract details table with brand colors
        contract_data = [
            ['Contract Name:', contract.contract_name],
            ['Upload Date:', contract.upload_date.strftime('%Y-%m-%d %H:%M') if contract.upload_date else 'N/A'],
            ['Status:', contract.status.value if contract.status else 'Unknown'],
            ['Total Clauses:', str(len(contract.clause_ids) if contract.clause_ids else 0)],
            ['Total Findings:', str(len(findings))],
        ]

        contract_table = Table(contract_data, colWidths=[2*inch, 4*inch])
        contract_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), BrandColors.VERY_LIGHT_GREY),
            ('TEXTCOLOR', (0, 0), (-1, -1), BrandColors.DARK_GREY),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), Typography.PRIMARY_BOLD),
            ('FONTNAME', (1, 0), (1, -1), Typography.PRIMARY_FONT),
            ('FONTSIZE', (0, 0), (-1, -1), Typography.BODY),
            ('GRID', (0, 0), (-1, -1), 0.5, BrandColors.LIGHT_GREY),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        story.append(contract_table)
        story.append(Spacer(1, 0.3*inch))

        # Findings summary by severity with brand colors
        severity_counts = self._count_by_severity(findings)

        subheading_style = ParagraphStyle(
            'Subheading',
            fontName=Typography.PRIMARY_BOLD,
            fontSize=Typography.HEADING_2,
            textColor=BrandColors.PRIMARY_BLUE,
            spaceAfter=Layout.PARAGRAPH_SPACING
        )

        story.append(Paragraph("Findings by Severity", subheading_style))
        story.append(Spacer(1, 0.1*inch))

        severity_data = [['Severity', 'Count', 'Percentage']]
        total = len(findings)

        for severity in [Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM, Severity.LOW]:
            count = severity_counts.get(severity, 0)
            pct = (count / total * 100) if total > 0 else 0
            severity_data.append([
                severity.value.capitalize(),
                str(count),
                f"{pct:.1f}%"
            ])

        severity_table = Table(severity_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        severity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BrandColors.PRIMARY_BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), BrandColors.WHITE),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), Typography.PRIMARY_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), Typography.BODY_LARGE),
            ('FONTNAME', (0, 1), (-1, -1), Typography.PRIMARY_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), Typography.BODY),
            ('GRID', (0, 0), (-1, -1), 0.5, BrandColors.LIGHT_GREY),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))

        story.append(severity_table)
        story.append(Spacer(1, 0.3*inch))

        # Total estimated impact with emphasis
        total_impact = sum(
            f.estimated_impact.value
            for f in findings
            if f.estimated_impact and f.estimated_impact.value
        )

        if total_impact > 0:
            impact_style = ParagraphStyle(
                'ImpactHighlight',
                fontName=Typography.PRIMARY_BOLD,
                fontSize=Typography.BODY_LARGE,
                textColor=BrandColors.DARK_GREY
            )
            impact_text = f"<b>Total Estimated Financial Impact:</b> ${total_impact:,.2f} USD"
            story.append(Paragraph(impact_text, impact_style))

        return story

    def _build_findings_section(
        self,
        findings: List[LeakageFinding],
        styles
    ) -> List:
        """Build findings detail section for PDF."""
        story = []

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12
        )

        story.append(Paragraph("Detailed Findings", heading_style))
        story.append(Spacer(1, 0.2*inch))

        # Sort by severity
        sorted_findings = sorted(
            findings,
            key=lambda f: (
                ['critical', 'high', 'medium', 'low'].index(f.severity.value.lower())
            )
        )

        for i, finding in enumerate(sorted_findings, 1):
            # Finding header
            severity_color = self._get_severity_color(finding.severity)
            finding_title = f"Finding #{i}: {finding.risk_type}"

            title_para = Paragraph(
                f"<b>{finding_title}</b>",
                ParagraphStyle(
                    'FindingTitle',
                    parent=styles['Heading3'],
                    fontSize=12,
                    textColor=severity_color,
                    spaceAfter=6
                )
            )
            story.append(title_para)

            # Finding details
            finding_data = [
                ['Category:', finding.leakage_category.value],
                ['Severity:', finding.severity.value.upper()],
                ['Confidence:', f"{finding.confidence * 100:.0f}%"],
                ['Detection Method:', finding.detection_method.value.upper()],
            ]

            if finding.estimated_impact and finding.estimated_impact.value:
                finding_data.append([
                    'Estimated Impact:',
                    f"${finding.estimated_impact.value:,.2f} {finding.estimated_impact.currency}"
                ])

            finding_table = Table(finding_data, colWidths=[2*inch, 4*inch])
            finding_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            story.append(finding_table)
            story.append(Spacer(1, 0.1*inch))

            # Explanation
            if finding.explanation:
                story.append(Paragraph("<b>Explanation:</b>", styles['Normal']))
                story.append(Paragraph(finding.explanation, styles['Normal']))
                story.append(Spacer(1, 0.1*inch))

            # Recommended action
            if finding.recommended_action:
                story.append(Paragraph("<b>Recommended Action:</b>", styles['Normal']))
                story.append(Paragraph(finding.recommended_action, styles['Normal']))

            story.append(Spacer(1, 0.2*inch))

        return story

    def _build_clauses_section(
        self,
        clauses: List[Clause],
        styles
    ) -> List:
        """Build clauses appendix section for PDF."""
        story = []

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12
        )

        story.append(Paragraph("Appendix: Extracted Clauses", heading_style))
        story.append(Spacer(1, 0.2*inch))

        for i, clause in enumerate(clauses[:50], 1):  # Limit to first 50
            clause_title = f"Clause {i} ({clause.clause_type})"
            story.append(Paragraph(f"<b>{clause_title}</b>", styles['Heading4']))
            story.append(Paragraph(clause.normalized_summary or clause.original_text[:500], styles['Normal']))
            story.append(Spacer(1, 0.15*inch))

        return story

    def _create_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        contract: Contract,
        findings: List[LeakageFinding]
    ):
        """Create summary worksheet in Excel."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Contract Leakage Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # Contract details
        row = 3
        ws[f'A{row}'] = "Contract Name:"
        ws[f'B{row}'] = contract.contract_name
        self._apply_header_style(ws[f'A{row}'])

        row += 1
        ws[f'A{row}'] = "Upload Date:"
        ws[f'B{row}'] = contract.upload_date.strftime('%Y-%m-%d %H:%M') if contract.upload_date else 'N/A'
        self._apply_header_style(ws[f'A{row}'])

        row += 1
        ws[f'A{row}'] = "Status:"
        ws[f'B{row}'] = contract.status.value if contract.status else 'Unknown'
        self._apply_header_style(ws[f'A{row}'])

        row += 1
        ws[f'A{row}'] = "Total Findings:"
        ws[f'B{row}'] = len(findings)
        self._apply_header_style(ws[f'A{row}'])

        # Severity summary
        row += 2
        ws[f'A{row}'] = "Findings by Severity"
        ws[f'A{row}'].font = Font(size=12, bold=True)

        row += 1
        headers = ['Severity', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self._apply_header_style(cell)

        severity_counts = self._count_by_severity(findings)
        total = len(findings)

        for severity in [Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM, Severity.LOW]:
            row += 1
            count = severity_counts.get(severity, 0)
            pct = (count / total * 100) if total > 0 else 0

            ws[f'A{row}'] = severity.value.capitalize()
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{pct:.1f}%"

        # Auto-size columns
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20

    def _create_findings_sheet(
        self,
        wb: openpyxl.Workbook,
        findings: List[LeakageFinding]
    ):
        """Create findings worksheet in Excel."""
        ws = wb.create_sheet("Findings")

        # Headers
        headers = [
            'Finding ID', 'Category', 'Severity', 'Risk Type',
            'Confidence', 'Detection Method', 'Estimated Impact',
            'Explanation', 'Recommended Action'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            self._apply_header_style(cell)

        # Data rows
        for row, finding in enumerate(findings, 2):
            ws.cell(row=row, column=1).value = finding.id
            ws.cell(row=row, column=2).value = finding.leakage_category.value
            ws.cell(row=row, column=3).value = finding.severity.value.upper()
            ws.cell(row=row, column=4).value = finding.risk_type
            ws.cell(row=row, column=5).value = f"{finding.confidence * 100:.0f}%"
            ws.cell(row=row, column=6).value = finding.detection_method.value.upper()

            if finding.estimated_impact and finding.estimated_impact.value:
                ws.cell(row=row, column=7).value = f"${finding.estimated_impact.value:,.2f}"
            else:
                ws.cell(row=row, column=7).value = "N/A"

            ws.cell(row=row, column=8).value = finding.explanation or ""
            ws.cell(row=row, column=9).value = finding.recommended_action or ""

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20

    def _create_clauses_sheet(
        self,
        wb: openpyxl.Workbook,
        clauses: List[Clause]
    ):
        """Create clauses worksheet in Excel."""
        ws = wb.create_sheet("Clauses")

        # Headers
        headers = [
            'Clause ID', 'Type', 'Section Number', 'Risk Signals',
            'Confidence', 'Summary'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            self._apply_header_style(cell)

        # Data rows
        for row, clause in enumerate(clauses, 2):
            ws.cell(row=row, column=1).value = clause.id
            ws.cell(row=row, column=2).value = clause.clause_type
            ws.cell(row=row, column=3).value = clause.section_number or "N/A"
            ws.cell(row=row, column=4).value = ", ".join(clause.risk_signals) if clause.risk_signals else "None"
            ws.cell(row=row, column=5).value = f"{clause.extraction_confidence * 100:.0f}%" if clause.extraction_confidence else "N/A"
            ws.cell(row=row, column=6).value = clause.normalized_summary or clause.original_text[:200]

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            if col == 6:  # Summary column
                ws.column_dimensions[column_letter].width = 50
            else:
                ws.column_dimensions[column_letter].width = 20

    def _apply_header_style(self, cell):
        """Apply header styling to Excel cell."""
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _count_by_severity(self, findings: List[LeakageFinding]) -> Dict[Severity, int]:
        """Count findings by severity."""
        counts = {}
        for finding in findings:
            counts[finding.severity] = counts.get(finding.severity, 0) + 1
        return counts

    def _get_severity_color(self, severity: Severity):
        """Get brand-consistent color for severity level."""
        severity_colors = {
            Severity.CRITICAL: BrandColors.CRITICAL_RED,
            Severity.HIGH: BrandColors.HIGH_ORANGE,
            Severity.MEDIUM: BrandColors.MEDIUM_YELLOW,
            Severity.LOW: BrandColors.LOW_GREEN,
        }
        return severity_colors.get(severity, BrandColors.BLACK)
